#!/usr/bin/env python3

import pandas as pd
import numpy as np
import seaborn as sb
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import statistics
import argparse
import dataclasses
# excel export with openpyxl
import openpyxl
# and xlsxwriter with pandas, though that could be modified to use openpyxl as well
import xlsxwriter
# for image saving
from io import BytesIO
# for date/time conversions
from datetime import datetime, timezone

# get summary statistics (min, max, range, mean, median, mode, and standard deviation for N50, sequence output, and flow cells per sample)
def get_summary_statistics(column):
    # define summary statistics class
    @dataclasses.dataclass
    class summary_statistics:
        total : int = 0
        min : float = 0
        max : float = 0
        range : float = 0
        mean : float = 0
        median : float = 0
        mode : float = 0
        stdev : float = 0
	# take single column of data frame as input
	# fill summary statistics class with each statistic
    # note that the line below should be used to give the total run count (across experiments and samples). Not useful with unique experiments, samples, etc.
    # as in flow cells per unique sample
    # later altered total to exclude NA values
    # summary_statistics.total = len(column)
    summary_statistics.total = column.count()
    # only calculate statistics if at least one non-NA element
    if summary_statistics.total > 0:
        # summary_statistics.min = min(column)
        summary_statistics.min = column.min()
        # summary_statistics.max = max(column)
        summary_statistics.max = column.max()
        summary_statistics.range = summary_statistics.max - summary_statistics.min
        # drop nas before calculating below if necessary
        # summary_statistics.mean = statistics.mean(column)
        # replaced statistics library calls with numpy calls
        # round to 3 decimal places
        summary_statistics.mean = round(np.nanmean(column),3)
        summary_statistics.median = round(np.nanmedian(column),3)
        summary_statistics.mode = round(statistics.mode(column),3)
        summary_statistics.stdev = round(np.nanstd(column),3)
    else:
        summary_statistics.min = np.nan
        summary_statistics.max = np.nan
        summary_statistics.range = np.nan
        summary_statistics.mean = np.nan
        summary_statistics.median = np.nan
        summary_statistics.mode = np.nan
        summary_statistics.stdev = np.nan
    # return data structure with each summary statistic as an attribute
    return summary_statistics
    
# get output per flow cell in two column list
def get_output_per_flow_cell(flow_cell_IDs, output, topup):
    # make data frame of flow cell IDs and output
    flow_cells_to_output = pd.concat([flow_cell_IDs, output, topup], axis=1, join='inner')
    # find unique flow cells
    unique_flow_cells = np.unique(flow_cell_IDs.astype(str))
    # create output_per_flow_cell_df data frame
    output_per_flow_cell_df = pd.DataFrame({'Flow Cell ID' : unique_flow_cells}, index=unique_flow_cells, columns=['Flow Cell ID','Flow cell output (Gb)','Run type'])
    # find flow cells and total output per unique experiment by iterating through unique experiment names
    for i in unique_flow_cells:
        unique_outputs_per_flow_cell = pd.unique(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Data output (Gb)'])
        unique_topup_per_flow_cell = pd.unique(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Run type'])
        # total_unique_outputs_per_flow_cell = len(unique_outputs_per_flow_cell)
        total_output_per_flow_cell = sum(flow_cells_to_output[flow_cells_to_output['Flow Cell ID'] == i]['Data output (Gb)'])
        output_per_flow_cell_df['Flow cell output (Gb)'][i] = total_output_per_flow_cell
        output_per_flow_cell_df['Run type'][i] = unique_topup_per_flow_cell
    # convert flow cell output total to numeric type for plotting
    output_per_flow_cell_df['Flow cell output (Gb)'] = pd.to_numeric(output_per_flow_cell_df['Flow cell output (Gb)'])
    # return flow_cells_per_experiment_df data frame
    return output_per_flow_cell_df
    
# get flow cells per experiment in two column list
def get_flow_cells_and_output_per_experiment(experiments, flow_cell_IDs, output):
    # take one column of experiments and one column of flow cell IDs from imported data frame as input
    # remove extraneous suffixes (e.g., "_topup") from experiment names
    # don't remove alphabetical characters altogether - e.g., PPMI_BLOOD_SSTEST is an experiment
    # experiment names should reflect independent brain isolates (e.g., PPMI_3080)
    experiments = experiments.str.replace(r'_topup', '', regex=True)
    experiments = experiments.str.replace(r'_recovery', '', regex=True)
    # dashes and underscores are the same thing so change all dashes to underscores
    experiments = experiments.str.replace(r'-', '_', regex=True)
    # make data frame of experiment names and flow cell IDs
    flow_cells_and_output_to_experiments = pd.concat([experiments, flow_cell_IDs, output], axis=1, join='inner')
    # find unique experiment names
    unique_experiments = np.unique(experiments)
    # initialized flow_cells_per_experiment list
    # create flow_cells_per_experiment_df data frame
    flow_cells_and_output_per_experiment_df = pd.DataFrame({'Experiment Name' : unique_experiments}, index=unique_experiments, columns=['Experiment Name','Flow Cells','Total output (Gb)'])
    # find flow cells and total output per unique experiment by iterating through unique experiment names
    for i in unique_experiments:
        unique_flow_cells_per_experiment = pd.unique(flow_cells_and_output_to_experiments[flow_cells_and_output_to_experiments['Experiment Name'] == i]['Flow Cell ID'])
        total_unique_flow_cells_per_experiment = len(unique_flow_cells_per_experiment)
        total_output_per_experiment = sum(flow_cells_and_output_to_experiments[flow_cells_and_output_to_experiments['Experiment Name'] == i]['Data output (Gb)'])
        flow_cells_and_output_per_experiment_df['Flow Cells'][i] = total_unique_flow_cells_per_experiment
        flow_cells_and_output_per_experiment_df['Total output (Gb)'][i] = total_output_per_experiment
    # convert output flow cell counts and total output totals to numeric type for plotting
    flow_cells_and_output_per_experiment_df['Flow Cells'] = pd.to_numeric(flow_cells_and_output_per_experiment_df['Flow Cells'])
    flow_cells_and_output_per_experiment_df['Total output (Gb)'] = pd.to_numeric(flow_cells_and_output_per_experiment_df['Total output (Gb)'])
    # return flow_cells_per_experiment_df data frame
    return flow_cells_and_output_per_experiment_df
    
# return flow cells per experiment distribution
# get total experiment count for each number of flow cells needed to complete experiment (approach 30x?)
def get_flow_cells_per_experiment_dist(column):
    # put distribution in flow_cells_per_experiment_dist
    # turn arange list into array and then 1d array with np.array() and flatten()
    flow_cells_per_experiment_dist = np.histogram(column, bins = np.array([np.arange(1,max(column)+2)]).flatten())
    # put output into labeled data frame for export
    flow_cells_per_experiment_dist_df = pd.DataFrame({'Flow Cells' : flow_cells_per_experiment_dist[1][0:max(column)], 'Frequency' : flow_cells_per_experiment_dist[0]}, columns=['Flow Cells', 'Frequency'])
    # return flow_cells_per_experiment_df data frame
    return flow_cells_per_experiment_dist_df

# get MinKNOW version distribution
def get_minknow_version_dist(column):
    # unique minknow versions
    unique_minknow_versions = np.unique(column)
    # create minknow_version_dist_df data frame
    minknow_version_dist_df = pd.DataFrame({'MinKNOW Version' : unique_minknow_versions}, index=unique_minknow_versions, columns=['MinKNOW Version', 'Frequency'])
    # count numbers of each in dataset
    for i in unique_minknow_versions:
        minknow_version_dist_df['Frequency'][i] = list(column).count(i)
    # return data frame with versions and counts per version
    return minknow_version_dist_df
    
# get sample rate distribution
def get_sample_rate_dist(column):
    # unique sample rates
    unique_sample_rates = np.unique(column)
    # create sample_rate_dist_df data frame
    sample_rate_dist_df = pd.DataFrame({'Sample Rate (Hz)' : unique_sample_rates}, index=unique_sample_rates, columns=['Sample Rate (Hz)', 'Frequency'])
    # count numbers of each in dataset
    for i in unique_sample_rates:
        sample_rate_dist_df.loc[i, 'Frequency'] = list(column).count(i)
    # return data frame with versions and counts per version
    return sample_rate_dist_df
    
# make summary statistic data frame
def make_summary_statistics_data_frame(summary_statistics_set, property_names):
    # set column names
    column_names = ['Property', 'Total', 'Min', 'Max', 'Mean', 'Median', 'Mode', 'Standard Deviation']
    # initialize data frame
    summary_statistics_df = pd.DataFrame(index=property_names, columns=column_names)
    # iterate through each property and populate data frame with summary_statistics_set attributes
    for idx, name in enumerate(property_names):
        summary_statistics_df.loc[name] = [property_names[idx],summary_statistics_set[idx].total,summary_statistics_set[idx].min,summary_statistics_set[idx].max,summary_statistics_set[idx].mean,summary_statistics_set[idx].median,summary_statistics_set[idx].mode,summary_statistics_set[idx].stdev]
    # return populated summary statistics data frame
    return summary_statistics_df
    
# identify topups and reconnections (flow cell moved and run restarted)
# later modified this to change "Initial run" designation to "Standard run" and "Interrupted"
def identify_topups(column):
    # make output array "topups" as long as input column
    topups = [None] * len(column)
    for idx, i in enumerate(column):
        if "topup" in i:
            # if topup in sample name, set value to topup
            topups[idx] = "Top up"
        elif "recovery" in i:
            # if recovery in sample name, set value to recovery
            topups[idx] = "Recovery"
        elif "reconnected" in i:
            # if reconnected in sample name, set value to reconnection
            topups[idx] = "Reconnection"
        else:
            # if topup or other labels not in sample name, set value to "Standard run"
            topups[idx] = "Standard run"
    # return topups/no topups column
    return topups

# identify reconnections through sequence run flow cell ID (shared between runs), date, and experiment name
# reconnection if same name and flow cell ID as previous run; add value to topups column shown before
# later modified this to change "Initial run" designation to "Standard run" and "Interrupted"
def identify_reconnections(data):
    # make copy of initial data to be modified further
    # sort data with reconnections by date and time so that first chronological run is "Interrupted" and subsequent are "Reconnection"
    data_with_reconnections = data.copy().sort_values(by='Start Run Timestamp')
    # get duplicate flow cells and remove NAs/NaNs
    duplicate_flow_cell_ids=data_with_reconnections[data_with_reconnections['Flow Cell ID'].duplicated()]['Flow Cell ID'].dropna()
    # loop through rows of data frame on duplicate flow cell ids
    for i in duplicate_flow_cell_ids:
        # get current duplicate flow cell ID data frame rows
        current_duplicate_flow_cell_id_df = data_with_reconnections[data_with_reconnections['Flow Cell ID'] == i]
        # get indexes to reference original df
        current_duplicate_flow_cell_id_indexes=current_duplicate_flow_cell_id_df.index
        # check if same sample name in all cases
        if (current_duplicate_flow_cell_id_df['Sample Name'].nunique() == 1):
            # set run type first instance to Interrupted [0]
            data_with_reconnections.loc[current_duplicate_flow_cell_id_indexes[0],'Run type'] = 'Interrupted'
            # set run type second and subsequent instances to Reconnection [1:]
            data_with_reconnections.loc[current_duplicate_flow_cell_id_indexes[1:],'Run type'] = 'Reconnection'
    # prior method left below as comments
    # loop through rows of data frame on Flow Cell ID column
    # for idx, i in enumerate(data['Flow Cell ID']):
        # see if flow cell ID in previous subset of list before current element
        # if i in list(data['Flow Cell ID'][0:idx]):
            # get data frame for repeated flow cell
            # duplicate_flow_cell_id_df = data[data['Flow Cell ID'] == i]
            # compare sample name for current run to all those with flow cell ID
            # test_sample_name = data.loc[idx]["Sample Name"]
            # print(test_sample_name)
            # if same sample name more than once for given flow cell, name most recent run as reconnection
            # if same sample name more than once for given flow cell, name first run as interrupted
            # name all subsequent runs as reconnection
            # if (sum(duplicate_flow_cell_id_df["Sample Name"] == test_sample_name) > 1):
                # change top up column to run type
                # data_with_reconnections.loc[(idx,"Run type")] = "Reconnection"
    # return data frame with algorithmically detected reconnections in topups column
    return data_with_reconnections
    
# make violinplot/swarmplot figure worksheet in output workbook
def make_violinswarmplot_worksheet(data,input_variable,group_variable,legend_patches,user_palette,strip_plot_set,workbook,worksheet_name,x_axis_title=None,cutoff=None,title=None,top_up=None):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make swarm plot to show how data points overlap with distribution
    # color points based on top up variable (initial run, top up, reconnection, recovery)
    # replace color='black'
    # set up plots differently depending on whether group variable is set
    if (group_variable is None):
        # make swarm plot to show how data points overlap with distribution
        # replace color='black'
        if strip_plot_set is False:
            if top_up is not None:
                rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
                ax = sb.swarmplot(data=data,x=input_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,edgecolor='white',linewidth=1)
            else:
                ax = sb.swarmplot(data=data,x=input_variable,color='black')
        elif strip_plot_set is True:
            if top_up is not None:
                rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
                ax = sb.stripplot(data=data,x=input_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,edgecolor='white',linewidth=1)
            else:
                ax = sb.stripplot(data=data,x=input_variable,color='black')
        # add violin plot using seaborn (sb.violinplot)
        # increase transparency to improve swarmplot visibility
        # use boxplot since only one "group" shown
        ax = sb.violinplot(data=data,x=input_variable,color='white',ax=ax)
    else:
        # make swarm plot to show how data points overlap with distribution
        # input variable on y axis and group on x axis
        # thus vertical swarm/violinplots instead of horizontal ones when no group specified
        # replace color='black' with hue set to group variable
        if strip_plot_set is False:
            # allow user set palette
            if user_palette is None:
                if top_up is not None:
                    rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
                    ax = sb.swarmplot(data=data,x=group_variable,y=input_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,edgecolor='white',linewidth=1)
                else:
                    ax = sb.swarmplot(data=data,x=group_variable,y=input_variable,hue=group_variable,legend=False)
            else:
                ax = sb.swarmplot(data=data,x=group_variable,y=input_variable,hue=group_variable,palette=user_palette,legend=False)
        elif strip_plot_set is True:
            # allow user set palette
            if user_palette is None:
                if top_up is not None:
                    rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
                    ax = sb.stripplot(data=data,x=group_variable,y=input_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,edgecolor='white',linewidth=1)
                else:
                    ax = sb.stripplot(data=data,x=group_variable,y=input_variable,hue=group_variable,legend=False)
            else:
                ax = sb.stripplot(data=data,x=group_variable,y=input_variable,hue=group_variable,palette=user_palette,legend=False)
        # add violin plot using seaborn (sb.violinplot)
        # increase transparency to improve swarmplot visibility
        # include quartile lines in this context (for easily, visually comparing between groups)
        ax = sb.violinplot(data=data,x=group_variable,y=input_variable,color='white',inner="quartile",ax=ax)
    # add x axis title if specified 
    if x_axis_title is not None:
        ax.set(xlabel=x_axis_title)
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # add red line for 90GB/30X cutoff or whatever necessary for specific plots
    # cutoff line defined by x value
    if cutoff is not None:
        # vertical line if non-grouped
        if (group_variable is None):
            ax.axvline(x=cutoff,color='red')
        # horizontal line if grouped (group variable on x-axis)
        else:
            ax.axhline(y=cutoff,color='red')
    # add legend with colors if requested
    if legend_patches is not None:
        plt.legend(handles=legend_patches)
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 200 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=200, bbox_inches='tight')
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    # close figure with matplotlib plt close
    plt.close()
    
# function for handling platform qc and calculating differences with starting active pores
def platform_qc_starting_active_pore_diff(data,platform_qc):
    # fix platform_qc column for joining
    platform_qc['Flow Cell ID'] = platform_qc['flow_cell_id']
    # join platform_qc and data tables on Flow Cell ID
    data_platform_qc_join = data.join(platform_qc.set_index("Flow Cell ID"),on="Flow Cell ID")
    # remove those where flow_cell_id is NaN
    data_platform_qc_join_cleaned = data_platform_qc_join.dropna(subset='flow_cell_id')
    # exclude NaN rows after join
    # find unique run timestamps
    unique_run_timestamps = np.unique(data_platform_qc_join_cleaned["Start Run Timestamp"])
    # initialize new data frame as long as unique runs based on unique run timestamps
    # remove {'Start Run Timestamp' : unique_run_timestamps} from arguments below
    data_with_platform_qc_and_diff = pd.DataFrame(columns=data_platform_qc_join_cleaned.columns.tolist() + ['Pore Difference','Time Difference'])
    # find matching rows in input platform qc table
    for idx, i in enumerate(unique_run_timestamps):
        # pick all platform qc results per run timestamp
        per_unique_run_df=data_platform_qc_join_cleaned[data_platform_qc_join_cleaned["Start Run Timestamp"] == i]
        # calculate difference between platform qc and starting active pores
        per_unique_run_df.loc[:,['Pore Difference']]=abs(per_unique_run_df['total_pore_count']-per_unique_run_df['Starting Active Pores'])
        # calculate difference between platfrom qc and starting active pore timestamps
        per_unique_run_df.loc[:,['Time Difference']]=abs(per_unique_run_df['timestamp']-per_unique_run_df['Start Run Timestamp'])
        # select per sample based on minimum time between run and platform qc check
        # get only first result if duplicates and add to table
        # below are some commands I tried and decided to ignore
        # data_with_platform_qc_and_diff.loc[idx,:]=per_unique_run_df[per_unique_run_df['Time Difference']==min(per_unique_run_df['Time Difference'])].iloc[0,:]
        # append deprecated in pandas 2.0
        # data_with_platform_qc_and_diff = data_with_platform_qc_and_diff.append(per_unique_run_df[per_unique_run_df['Time Difference']==min(per_unique_run_df['Time Difference'])].iloc[0,:])
        # below for debugging
        # print(per_unique_run_df)
        data_with_platform_qc_and_diff.loc[idx] = per_unique_run_df[per_unique_run_df['Time Difference']==min(per_unique_run_df['Time Difference'])].iloc[0,:]
    # set dtypes to match original joined df
    data_with_platform_qc_and_diff=data_with_platform_qc_and_diff.astype(data_platform_qc_join_cleaned.dtypes)
    # convert starting active pore column to numeric
    # data_with_platform_qc_and_diff.loc[:,['Starting Active Pores']]=pd.to_numeric(data_with_platform_qc_and_diff['Starting Active Pores'])
    # return data frame with platform qc active pores, pore differences, and timestamp differences appended
    return data_with_platform_qc_and_diff

# function for calculating storage time based on imported delivery date table
def calc_storage_time_from_delivery_date(data,delivery_date_df):
    # join delivery date and data tables on Flow Cell ID
    data_delivery_date_join_df = data.join(delivery_date_df.set_index("Flow Cell ID"),on="Flow Cell ID")
    # remove those where Batch is NaN
    data_delivery_date_join_df = data_delivery_date_join_df.dropna(subset='Batch')
    # calculate storage time
    # data_delivery_date_join_df.loc[:,['Storage Time (Days)']] = round((data_delivery_date_join_df['Start Run Timestamp'] - data_delivery_date_join_df['Delivery date timestamp'])/86400,3)
    # find unique run timestamps
    unique_run_timestamps = np.unique(data_delivery_date_join_df["Start Run Timestamp"])
    # initialize new data frame as long as unique runs based on unique run timestamps
    # remove {'Start Run Timestamp' : unique_run_timestamps} from arguments below
    # convert joined table column names to list
    data_with_delivery_date_batch_and_storage_time = pd.DataFrame(columns=data_delivery_date_join_df.columns.tolist() + ['Storage Time (Days)'])
    # find matching rows in input delivery date/batch table
    for idx, i in enumerate(unique_run_timestamps):
        # pick all platform qc results per run timestamp
        per_unique_run_df=data_delivery_date_join_df[data_delivery_date_join_df["Start Run Timestamp"] == i]
        # calculate storage time
        per_unique_run_df.loc[:,['Storage Time (Days)']] = (per_unique_run_df['Start Run Timestamp'] - per_unique_run_df['Delivery date timestamp'])/86400
        # if multiple delivery dates for unique run, then use most recent (shortest storage time calculation)
        data_with_delivery_date_batch_and_storage_time.loc[idx] = per_unique_run_df[per_unique_run_df['Storage Time (Days)']==min(per_unique_run_df['Storage Time (Days)'])].iloc[0,:]
    # set dtypes to match original joined df
    data_with_delivery_date_batch_and_storage_time=data_with_delivery_date_batch_and_storage_time.astype(data_delivery_date_join_df.dtypes)
    # return data frame with delivery date/batch info and storage times calculated as differences between start run timestamps and delivery date timestamps
    return data_with_delivery_date_batch_and_storage_time

# single function for scatterplots with/without cutoffs

def make_scatterplot_worksheet(data,group_variable,legend_patches,user_palette,strip_plot_set,workbook,worksheet_name,title=None,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable=None,y_variable=None,prop_point_size=False,size_column=None,has_date_time=False):
    # create worksheet for figure output
    worksheet=workbook.create_sheet(worksheet_name)
    # initialize raw data buffer for image
    imgdata=BytesIO()
    # initialize plot overall
    fig, ax = plt.subplots()
    # make scatterplot of active pores vs. per flow cell data output
    # include regression by using sb.regplot() function if show_reg_line=True
    # had to remove regression to use hue keyword
    # color points by topup/not topup run if show_run_colors is True
    if show_run_colors is True and prop_point_size is False:
        # show top up colors if no group variable included
        if group_variable is None:
            rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
            ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette)
        # override top up colors if group variable included
        else:
            # default palette if no palette specified
            if user_palette is None:
                ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue=group_variable)
            # otherwise use user specified palette
            else:
                ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue=group_variable,palette=user_palette)
    # add point size if specified (prop_point_size is True)
    # note use of data[size_column] as point size
    elif show_run_colors is True and prop_point_size is True:
        # show top up colors if no group variable included
        if group_variable is None:
            rearranged_color_palette = [sb.color_palette()[0],'firebrick',sb.color_palette()[1],sb.color_palette()[4],sb.color_palette()[5]]
            ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue="Run type",hue_order=['Standard run','Interrupted','Top up','Reconnection','Recovery'],palette=rearranged_color_palette,size=data[size_column])
        # override top up colors if no group variable included
        else:
            # default palette if no palette specified
            if user_palette is None:
                ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue=group_variable,size=data[size_column])
            # otherwise use user specified palette
            else:
                ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,hue=group_variable,palette=user_palette,size=data[size_column])
    elif show_run_colors is False and prop_point_size is False:
        ax = sb.scatterplot(data=data,x=x_variable,y=y_variable)
    # add point size if specified (prop_point_size is True)
    elif show_run_colors is False and prop_point_size is True:
        ax = sb.scatterplot(data=data,x=x_variable,y=y_variable,size=data[size_column])
    # show regression line if specified (show_reg_line is True)
    if show_reg_line is True and prop_point_size is False:
        ax = sb.regplot(data=data,x=x_variable,y=y_variable)
    # add point size if specified (prop_point_size is True)
    elif show_reg_line is True and prop_point_size is True:
        ax = sb.regplot(data=data,x=x_variable,y=y_variable,size=data[size_column])
    # add title if specified
    if title is not None:
        ax.set_title(title)
    # set minimum y and x to zero
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)
    if x_cutoffs is not None:
        for idx, i in enumerate(x_cutoffs):
            # add vertical x cutoffs sequentially
            ax.axvline(x=i,color=x_cutoff_colors[idx])
    if y_cutoffs is not None:
        for idx, i in enumerate(y_cutoffs):
            # add horizontal y cutoffs sequentially
            ax.axhline(y=i,color=y_cutoff_colors[idx])
    # add legend with colors if requested
    if legend_patches is not None:
        plt.legend(handles=legend_patches)
    # handling datetime based x axis
    if has_date_time is True:
        ax.set_xlim(data[x_variable].min(),data[x_variable].max())
        ax.tick_params(axis='x', rotation=45)
    # put figure in variable to prep for saving into buffer
    # fig = swarmplot.get_figure()
    # save figure as 150 dpi PNG into buffer
    fig.savefig(imgdata, format='png', dpi=150, bbox_inches='tight')
    # close figure
    fig.clf()
    # make openpyxl image from raw data
    img = openpyxl.drawing.image.Image(imgdata)
    # set location of image in worksheet (A1)
    img.anchor = 'A1'
    # add image to worksheet
    worksheet.add_image(img)
    # close figure with matplotlib plt close
    plt.close()
    
# set up command line argument parser
parser = argparse.ArgumentParser(description='This program gets summary statistics from long read sequencing report data.')

# get input and output arguments
# allow multiple inputs
parser.add_argument('-input', action="store", dest="input_file", nargs="+", help="Input tab-delimited tsv file(s) containing features extracted from long read sequencing reports.")
# if multiple inputs, require input names
parser.add_argument('-names', action="store", dest="names", nargs="*", help="Names corresponding to input tsv file(s); required if more than one tsv provided.")
# single output xlsx
parser.add_argument('-output', action="store", dest="output_file", help="Output long read sequencing summary statistics XLSX")
# import platform QC table in format specified in CARDlongread_MinKNOW_api_scripts repository
parser.add_argument('-platform_qc', action="store",default=None, dest="platform_qc", help="Input platform QC table to calculate active pore dropoff upon sequencing (optional)")
# import delivery date table parsed from ONT provided spreadsheet with export_flow_cell_delivery_dates.py
parser.add_argument('-delivery_date_batches',action="store",default=None,dest="delivery_date_batches", help="Input delivery date/batch table to calculate storage time per flow cell based on delivery date and run timestamp (optional).")
# add option to specify plot title
parser.add_argument('-plot_title', action="store", default=None, dest="plot_title", help="Title for each plot in output XLSX (optional)")
# add boolean --plot_cutoff argument
parser.add_argument('--plot_cutoff', action=argparse.BooleanOptionalAction, default=True, dest="plot_cutoff", help="Include cutoff lines in violin plots (optional; default true; --no-plot_cutoff to override)")
# include failed run cutoff to exclude as well
parser.add_argument('-run_cutoff', action="store", default=1, type=float, dest="run_cutoff", help="Minimum data output per flow cell run to include (optional, 1 Gb default)")
# add option for stripplot instead of swarmplot (in case of excessive data points)
parser.add_argument('--strip_plot', action=argparse.BooleanOptionalAction, default=False, dest="strip_plot", help="Show strip plots instead of swarm plots inside violin plots (optional; default false)")
# add option for color palette
parser.add_argument('-colors', action="store", default=None, dest="colors", nargs="*", help="Color palette corresponding to sequential groups displayed (e.g., 'blue', 'red', 'blue'); optional and used only if more than one tsv provided.")
# add option for custom legend colors
parser.add_argument('-legend_colors', action="store", default=None, dest="legend_colors", nargs="*", help="Colors shown in the legend (e.g., 'blue', 'red'); optional and used only if more color palette included above. Must be palette subset.")
# add option for custom legend labels
parser.add_argument('-legend_labels', action="store", default=None, dest="legend_labels", nargs="*", help="Labels for each color in legend in order specified in -legend_colors.")
# add option to show sample size for groups in grouped violinplots
parser.add_argument('--group_count', action=argparse.BooleanOptionalAction, default=False, dest="show_group_count", help="Show group count in x-axis labels (optional; default false)")
# add option to output platform qc joined table
parser.add_argument('-output_table_with_platform_qc', action="store", default=None, help="Output filename for run report summary table joined with platform QC flow cell check information (optional).")
# add option to output run type designation
parser.add_argument('-output_table_with_run_type', action="store", default=None, help="Output filename for run report summary table with appended run type, such as 'top up' or 'reconnection' (optional).")
# add option to output storage time designation
parser.add_argument('-output_table_with_storage_time', action="store", default=None, help="Output filename for run report summary table with delivery date, batch, and storage times in days added (optional).")

# parse arguments
results = parser.parse_args()

# throw error if no input file provided
if results.input_file is None:
	quit('ERROR: No input file (-input) provided!')
    
# throw error if no names provided if multiple input files provided
if len(results.input_file)>1:
    if len(results.names)<=1:
        quit('ERROR: Multiple input files provided but not multiple names (-names).')
        
# throw error if no names provided if multiple input files provided
if len(results.input_file)>1:
    if len(results.names)<=1:
        quit('ERROR: Multiple input files provided but not multiple names (-names).')
    elif len(results.names) != len(results.input_file):
        quit('ERROR: Number of names is different from number of input files.')
    # test if number of colors different from input files and names
    elif (results.colors is not None) and (len(results.colors) != len(results.names)):
        quit("ERROR: Color palette provided, but number of colors doesnt match number of group names.")
        
# set legend_patches to None by default
legend_patches=None
# test if legend colors and labels have proper length
if (results.legend_colors is not None) or (results.legend_labels is not None):
    if len(results.legend_colors) != len(results.legend_labels):
        quit('ERROR: Number of legend colors does not match number of legend labels.')
    else:
        # prepare legend patches
        # make list as long as legend colors (at this point same as legend_labels)
        legend_patches = [0] * len(results.legend_colors)
        for idx, i in enumerate(results.legend_colors):
            legend_patches[idx] = mpatches.Patch(color=i, label=results.legend_labels[idx])
        
# test if legend colors provided but not labels or vice versa
if ((results.legend_colors is not None) and (results.legend_labels is None)) or ((results.legend_colors is None) and (results.legend_labels is not None)):
    quit('ERROR: Either legend colors or legend labels provided but not both.')
        
# set default output filename
if results.output_file is None:
    results.output_file='output_summary_statistics.xlsx'

# read tab delimited output into pandas data frame
# case if just one input file provided
if len(results.input_file)==1:
    longread_extract_initial=pd.read_csv(results.input_file[0],sep='\t')
    # first filter out low output runs
    longread_extract = longread_extract_initial[longread_extract_initial['Data output (Gb)'] > results.run_cutoff]
    # fix indices
    longread_extract.reset_index(drop='True',inplace=True)
    # add top up column to data frame
    # avoid nested tuple warning
    # longread_extract["Top up"] = identify_topups(longread_extract["Sample Name"])
    # add after 12th column or last column (dataframe.shape[1])
    longread_extract.insert(longread_extract.shape[1],"Run type",identify_topups(longread_extract["Sample Name"]),True)
    # identify reconnections amongst flow cells
    longread_extract = identify_reconnections(longread_extract)
    # convert run starting timestamp to date and time
    longread_extract['Run date']=[datetime.fromtimestamp(x) for x in pd.to_numeric(longread_extract['Start Run Timestamp'])]
    # get flow cells/output per experiment table overall
    longread_extract_flow_cells_and_output_per_experiment = get_flow_cells_and_output_per_experiment(longread_extract['Experiment Name'], longread_extract['Flow Cell ID'], longread_extract['Data output (Gb)'])
    # get output per flow cell table overall
    longread_extract_output_per_flow_cell = get_output_per_flow_cell(longread_extract['Flow Cell ID'], longread_extract['Data output (Gb)'], longread_extract['Run type'])
    # set grouped variable as False
    grouped=False
    # output table with run type determined if specified in options
    if results.output_table_with_run_type is not None:
        # output to TSV with indexes excluded
        longread_extract.to_csv(results.output_table_with_run_type,index=False,sep="\t")
    # handle delivery dates
    if results.delivery_date_batches is not None:
        delivery_date_batches_table=pd.read_csv(results.delivery_date_batches,sep="\t")
        # set longread_extract to original table joined with delivery date/batches table + calculated storage times in days
        longread_extract=calc_storage_time_from_delivery_date(longread_extract,delivery_date_batches_table)
        # convert critical variable to numeric for statistics and plotting
        longread_extract['Storage Time (Days)']=pd.to_numeric(longread_extract['Storage Time (Days)'])
        # output table with delivery date determined if specified in options
        if results.output_table_with_storage_time is not None:
            longread_extract.to_csv(results.output_table_with_storage_time,index=False,sep="\t")
    
# what if multiple input files provided
elif len(results.input_file)>1:
    # store input tables in list as long input filename set
    longread_extract_initial_list=[0] * len(results.input_file)
    # store flow cells/output per experiment tables for each group
    longread_extract_flow_cells_and_output_per_experiment_initial_list=[0] * len(results.input_file)
    # store output per flow cell tables for each group
    longread_extract_output_per_flow_cell_initial_list=[0] * len(results.input_file)
    # iterate through groups
    for idx, i in enumerate(results.input_file): 
        longread_extract_initial_list[idx]=pd.read_csv(i,sep='\t')
        # first filter out low output runs
        longread_extract_initial_list[idx]=longread_extract_initial_list[idx][longread_extract_initial_list[idx]['Data output (Gb)'] > results.run_cutoff]
        # fix indices
        longread_extract_initial_list[idx].reset_index(drop='True',inplace=True)
        # add top up column to data frame
        # avoid nested tuple warning
        # longread_extract["Top up"] = identify_topups(longread_extract["Sample Name"])
        # add after 12th column or last column (dataframe.shape[1])
        longread_extract_initial_list[idx].insert(longread_extract_initial_list[idx].shape[1],"Run type",identify_topups(longread_extract_initial_list[idx]["Sample Name"]),True)
        # identify reconnections amongst flow cells
        longread_extract_initial_list[idx] = identify_reconnections(longread_extract_initial_list[idx])
        # convert run starting timestamp to date and time
        longread_extract_initial_list[idx]['Run date']=[datetime.fromtimestamp(x) for x in pd.to_numeric(longread_extract_initial_list[idx]['Start Run Timestamp'])]
        # get flow cells/output per experiment table for group
        longread_extract_flow_cells_and_output_per_experiment_initial_list[idx] = get_flow_cells_and_output_per_experiment(longread_extract_initial_list[idx]['Experiment Name'], longread_extract_initial_list[idx]['Flow Cell ID'], longread_extract_initial_list[idx]['Data output (Gb)'])
        # get output per flow cell table for group
        longread_extract_output_per_flow_cell_initial_list[idx] = get_output_per_flow_cell(longread_extract_initial_list[idx]['Flow Cell ID'], longread_extract_initial_list[idx]['Data output (Gb)'], longread_extract_initial_list[idx]['Run type'])
        # add group name to each table in list
        if results.show_group_count is True:
            # if group count specified, add group count to group name
            group_count = len(longread_extract_initial_list[idx])
            # in this way, show n=717 or similar below group names in all plots
            # note this is count of runs per group after filtering for per run cutoff
            longread_extract_initial_list[idx]['Group']=results.names[idx]
            longread_extract_initial_list[idx]['Group and count']=results.names[idx] + "\nn=" + str(group_count)
            # do this for additional distributions
            # change group count to match next data frame size
            # if group count specified, add group count to group name
            group_count = len(longread_extract_flow_cells_and_output_per_experiment_initial_list[idx])
            longread_extract_flow_cells_and_output_per_experiment_initial_list[idx]['Group']=results.names[idx]
            longread_extract_flow_cells_and_output_per_experiment_initial_list[idx]['Group and count']=results.names[idx] + "\nn=" + str(group_count)
            # repeat for output per flow cell table
            group_count = len(longread_extract_output_per_flow_cell_initial_list[idx])
            longread_extract_output_per_flow_cell_initial_list[idx]['Group']=results.names[idx]
            longread_extract_output_per_flow_cell_initial_list[idx]['Group and count']=results.names[idx] + "\nn=" + str(group_count)
            # group_names_list[idx]=results.names[idx] + "\nn=" + str(group_count)
        else:
            longread_extract_initial_list[idx]['Group']=results.names[idx]
            # add group name to additional distributions
            longread_extract_flow_cells_and_output_per_experiment_initial_list[idx]['Group']=results.names[idx]
            longread_extract_output_per_flow_cell_initial_list[idx]['Group']=results.names[idx]
    # combine groups into single concatenated data table
    longread_extract=pd.concat(longread_extract_initial_list[:],ignore_index=True)
    longread_extract_flow_cells_and_output_per_experiment=pd.concat(longread_extract_flow_cells_and_output_per_experiment_initial_list[:],ignore_index=True)
    longread_extract_output_per_flow_cell=pd.concat(longread_extract_output_per_flow_cell_initial_list[:],ignore_index=True)
    # set grouped variable as True
    grouped=True
    # output table with run type determined if specified in options
    if results.output_table_with_run_type is not None:
        # output to TSV with indexes excluded
        longread_extract.to_csv(results.output_table_with_run_type,index=False,sep="\t")
    # handle delivery dates
    if results.delivery_date_batches is not None:
        delivery_date_batches_table=pd.read_csv(results.delivery_date_batches,sep="\t")
        # set longread_extract to original table joined with delivery date/batches table + calculated storage times in days
        longread_extract=calc_storage_time_from_delivery_date(longread_extract,delivery_date_batches_table)
        # convert critical variables to numeric for statistics and plotting
        longread_extract['Storage Time (Days)']=pd.to_numeric(longread_extract['Storage Time (Days)'])
        # output table with delivery date determined if specified in options
        if results.output_table_with_storage_time is not None:
            longread_extract.to_csv(results.output_table_with_storage_time,index=False,sep="\t")
    
# read csv delimited platform qc file into pandas data frame if provided
if results.platform_qc is not None:
    platform_qc_table=pd.read_csv(results.platform_qc)

# use functions above
# if platform qc file provided, make joined longread_extract/platform_qc_table
if results.platform_qc is not None:
    longread_extract_with_platform_qc_and_diff=platform_qc_starting_active_pore_diff(longread_extract,platform_qc_table)
    # make Platform QC active pores column for plotting
    longread_extract_with_platform_qc_and_diff.loc[:,['Platform QC active pores']]=pd.to_numeric(longread_extract_with_platform_qc_and_diff['total_pore_count'])
    # convert all plotted columns to numeric
    longread_extract_with_platform_qc_and_diff['Pore Difference']=pd.to_numeric(longread_extract_with_platform_qc_and_diff['Pore Difference'])
    longread_extract_with_platform_qc_and_diff['Time Difference']=pd.to_numeric(longread_extract_with_platform_qc_and_diff['Time Difference'])
    longread_extract_with_platform_qc_and_diff['N50 (kb)']=pd.to_numeric(longread_extract_with_platform_qc_and_diff['N50 (kb)'])
    longread_extract_with_platform_qc_and_diff['Data output (Gb)']=pd.to_numeric(longread_extract_with_platform_qc_and_diff['Data output (Gb)'])
    # also include time series for plotting
    # convert to iso8601 format
    longread_extract_with_platform_qc_and_diff['Platform QC date']=[datetime.fromtimestamp(x) for x in pd.to_numeric(longread_extract_with_platform_qc_and_diff['timestamp'])]
    longread_extract_with_platform_qc_and_diff['Run date']=[datetime.fromtimestamp(x) for x in pd.to_numeric(longread_extract_with_platform_qc_and_diff['Start Run Timestamp'])] 
    # output longread_extract/platform_qc joined table if option specified
    # output table with platform QC stats joined if specified in options
    if results.output_table_with_platform_qc is not None:
        # output to TSV with indexes excluded
        longread_extract_with_platform_qc_and_diff.to_csv(results.output_table_with_platform_qc,index=False,sep="\t")       

# functionalize all below to run through on separate groups
def longread_platform_qc_summary_statistics(longread_extract,longread_extract_with_platform_qc_and_diff,longread_extract_flow_cells_and_output_per_experiment,longread_extract_output_per_flow_cell):
    # flow cells per experiment distribution
    longread_extract_flow_cells_per_experiment_dist = get_flow_cells_per_experiment_dist(longread_extract_flow_cells_and_output_per_experiment['Flow Cells'])
    # summary statistics on...
    # read N50
    read_N50_summary_stats = get_summary_statistics(longread_extract['N50 (kb)'])
    # sequence output
    sequence_output_summary_stats = get_summary_statistics(longread_extract['Data output (Gb)'])
    # read count per run
    read_count_summary_stats = get_summary_statistics(longread_extract['Read Count (M)'])
    # starting active pores per run
    starting_active_pores_summary_stats = get_summary_statistics(longread_extract['Starting Active Pores'])
    # platform QC active pores per run if possible
    if longread_extract_with_platform_qc_and_diff is not None:
        platform_qc_summary_stats = get_summary_statistics(longread_extract_with_platform_qc_and_diff['total_pore_count'])
        # differences between platform QC and starting active pores if possible
        pore_difference_qc_summary_stats = get_summary_statistics(longread_extract_with_platform_qc_and_diff['Pore Difference'])
    # flow cells per experiment
    flow_cells_per_experiment_summary_stats = get_summary_statistics(longread_extract_flow_cells_and_output_per_experiment['Flow Cells'])
    # output per flow cell
    output_per_flow_cell_summary_stats = get_summary_statistics(longread_extract_output_per_flow_cell['Flow cell output (Gb)'])
    # total output per experiment
    output_per_experiment_summary_stats = get_summary_statistics(longread_extract_flow_cells_and_output_per_experiment['Total output (Gb)'])
    # additional summary stats added 10/8/2025
    # active pore AUC
    active_pore_auc_summary_stats = get_summary_statistics(longread_extract['Active Pore AUC'])
    # average active pores
    average_active_pores_summary_stats = get_summary_statistics(longread_extract['Average Active Pores'])
    # starting pore occupancy
    starting_pore_occupancy_summary_stats = get_summary_statistics(longread_extract['Starting Pore Occupancy'])
    # average pore occupancy
    average_pore_occupancy_summary_stats = get_summary_statistics(longread_extract['Average Pore Occupancy'])
    # passed modal Q score
    passed_modal_q_score_summary_stats = get_summary_statistics(longread_extract['Passed Modal Q Score'])
    # failed modal Q score
    failed_modal_q_score_summary_stats = get_summary_statistics(longread_extract['Failed Modal Q Score'])
    # starting translocation speed
    starting_translocation_speed_summary_stats = get_summary_statistics(longread_extract['Starting Median Translocation Speed'])
    # average translocation speed
    average_translocation_speed_summary_stats = get_summary_statistics(longread_extract['Average Median Translocation Speed Over Time'])
    # starting median Q score
    starting_median_q_score_summary_stats = get_summary_statistics(longread_extract['Starting Median Q Score'])
    # average median Q score
    average_median_q_score_summary_stats = get_summary_statistics(longread_extract['Average Median Q Score Over Time'])
    # passed bases
    passed_bases_summary_stats = get_summary_statistics(longread_extract['Passed Bases (Gb)'])
    # failed bases
    failed_bases_summary_stats = get_summary_statistics(longread_extract['Failed Bases (Gb)'])
    # percentage passed bases
    percentage_passed_bases_summary_stats = get_summary_statistics(longread_extract['Percentage Passed Bases'])
    # check if storage time included in longread_extract
    if 'Storage Time (Days)' in longread_extract:
        storage_time_summary_stats = get_summary_statistics(longread_extract['Storage Time (Days)'])
    
    # combine summary stats into one list
    combined_summary_stats = [read_N50_summary_stats,
    sequence_output_summary_stats,
    read_count_summary_stats,
    starting_active_pores_summary_stats,
    average_active_pores_summary_stats,
    flow_cells_per_experiment_summary_stats,
    output_per_flow_cell_summary_stats,
    output_per_experiment_summary_stats,
    active_pore_auc_summary_stats,
    starting_pore_occupancy_summary_stats,
    average_pore_occupancy_summary_stats,
    passed_modal_q_score_summary_stats,
    failed_modal_q_score_summary_stats,
    starting_translocation_speed_summary_stats,
    average_translocation_speed_summary_stats,
    starting_median_q_score_summary_stats,
    average_median_q_score_summary_stats,
    passed_bases_summary_stats,
    failed_bases_summary_stats,
    percentage_passed_bases_summary_stats]

    # make data frame from combined_summary_stats
    combined_property_names = ['Read N50 (kb)',
    'Run data output (Gb)',
    'Run read count (millions)',
    'Starting active pores',
    'Average active pores',
    'Flow cells per experiment',
    'Flow cell output (Gb)', 
    'Total experiment output (Gb)',
    'Active pore AUC',
    'Starting pore occupancy',
    'Average pore occupancy',
    'Passed modal Q score',
    'Failed modal Q score',
    'Starting translocation speed',
    'Average translocation speed',
    'Starting median Q score',
    'Average median Q score',
    'Passed bases (Gb)',
    'Failed bases (Gb)',
    'Percentage passed bases']
    
    # append storage time to combined_summary_stats and combined_property_names if present:
    if 'Storage Time (Days)' in longread_extract:
        combined_summary_stats.append(storage_time_summary_stats)
        combined_property_names.append('Storage time (days)')
    
    combined_summary_stats_df = make_summary_statistics_data_frame(combined_summary_stats,combined_property_names)

    # include platform QC active pores/pore difference information where applicable
    if longread_extract_with_platform_qc_and_diff is not None:
        # combine summary stats into one list
        combined_summary_stats = [read_N50_summary_stats,
        sequence_output_summary_stats,
        read_count_summary_stats,
        starting_active_pores_summary_stats,
        average_active_pores_summary_stats,
        platform_qc_summary_stats,
        pore_difference_qc_summary_stats,
        flow_cells_per_experiment_summary_stats,
        output_per_flow_cell_summary_stats,
        output_per_experiment_summary_stats,
        active_pore_auc_summary_stats,
        starting_pore_occupancy_summary_stats,
        average_pore_occupancy_summary_stats,
        passed_modal_q_score_summary_stats,
        failed_modal_q_score_summary_stats,
        starting_translocation_speed_summary_stats,
        average_translocation_speed_summary_stats,
        starting_median_q_score_summary_stats,
        average_median_q_score_summary_stats,
        passed_bases_summary_stats,
        failed_bases_summary_stats,
        percentage_passed_bases_summary_stats]
        # make data frame from combined_summary_stats
        combined_property_names = ['Read N50 (kb)',
        'Run data output (Gb)',
        'Run read count (millions)',
        'Starting active pores',
        'Average active pores',
        'Platform QC active pores',
        'Pore difference',
        'Flow cells per experiment',
        'Flow cell output (Gb)',
        'Total experiment output (Gb)',
        'Active pore AUC',
        'Starting pore occupancy',
        'Average pore occupancy',
        'Passed modal Q score',
        'Failed modal Q score',
        'Starting translocation speed',
        'Average translocation speed',
        'Starting median Q score',
        'Average median Q score',
        'Passed bases (Gb)',
        'Failed bases (Gb)',
        'Percentage passed bases']
        
        # append storage time to combined_summary_stats and combined_property_names if present:
        if 'Storage Time (Days)' in longread_extract:
            combined_summary_stats.append(storage_time_summary_stats)
            combined_property_names.append('Storage time (days)')
        
        combined_summary_stats_df = make_summary_statistics_data_frame(combined_summary_stats,combined_property_names)
        
    # minknow version distribution
    longread_extract_minknow_version_dist = get_minknow_version_dist(longread_extract['MinKNOW Version'])
    # sampling rate distribution
    longread_extract_sample_rate_dist = get_sample_rate_dist(longread_extract['Sample Rate (Hz)'])
    # stop functionalizing here
    return(combined_summary_stats_df,longread_extract_minknow_version_dist,longread_extract_sample_rate_dist,longread_extract_flow_cells_per_experiment_dist)

# save data frames as tab-delimited file (.tsv)
# Example data structure
# Header 
# Property   Total   Min Max Mean    Median  Mode    Standard Deviation
# Read N50 (kb)
# Data output (Gb)
# Flow cells per experiment
# <empty>
# Flow Cells    Frequency
# 1 128
# 2 398
# etc.
# <empty>
# MinKNOW Version   Frequency
# 22.10.7   756
# 23.11.4   9
# etc.
# <empty>

if grouped is False:
    # run above summary statistics function
    if results.platform_qc is not None:
        (combined_summary_stats_df,longread_extract_minknow_version_dist,longread_extract_sample_rate_dist,longread_extract_flow_cells_per_experiment_dist)=longread_platform_qc_summary_statistics(longread_extract,longread_extract_with_platform_qc_and_diff,longread_extract_flow_cells_and_output_per_experiment,longread_extract_output_per_flow_cell) 
    else:
        (combined_summary_stats_df,longread_extract_minknow_version_dist,longread_extract_sample_rate_dist,longread_extract_flow_cells_per_experiment_dist)=longread_platform_qc_summary_statistics(longread_extract,None,longread_extract_flow_cells_and_output_per_experiment,longread_extract_output_per_flow_cell)
    # output data frames and figures to excel spreadsheet
    writer = pd.ExcelWriter(results.output_file)
    # write data frames with a row between each
    # write combined summary stats
    start_row = 0
    combined_summary_stats_df.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
    # write flow cells per experiment distribution
    # add 1 to row number after combined_summary_stats_df exported
    start_row = start_row + len(combined_summary_stats_df) + 2
    longread_extract_flow_cells_per_experiment_dist.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
    # write minknow version distribution
    start_row = start_row + len(longread_extract_flow_cells_per_experiment_dist) + 2
    longread_extract_minknow_version_dist.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
    # write sample rate distribution
    start_row = start_row + len(longread_extract_minknow_version_dist) + 2
    longread_extract_sample_rate_dist.to_excel(writer, startrow=start_row, index=False, sheet_name='Summary statistics report')
    # write flow cells and output per flow cell on another worksheet
    longread_extract_output_per_flow_cell.to_excel(writer, index=False, sheet_name='Output per flow cell ID')
    # write flow cells and output per unique experiment on another worksheet
    longread_extract_flow_cells_and_output_per_experiment.to_excel(writer, index=False, sheet_name='FC + output per experiment')
    # eventually write joined platform QC/summary table to worksheet (to do)
    # close writer and save workbook
    writer.close()
elif grouped is True:
    # output data frames and figures to excel spreadsheet
    writer = pd.ExcelWriter(results.output_file)
    # run through loop here
    # loop through all group names from input
    for idx, i in enumerate(results.names):
        # run above summary statistics function
        if results.platform_qc is not None:
            (combined_summary_stats_df,longread_extract_minknow_version_dist,longread_extract_sample_rate_dist,longread_extract_flow_cells_per_experiment_dist)=longread_platform_qc_summary_statistics(longread_extract[longread_extract['Group'] == i],longread_extract_with_platform_qc_and_diff[longread_extract_with_platform_qc_and_diff['Group']==i],longread_extract_flow_cells_and_output_per_experiment[longread_extract_flow_cells_and_output_per_experiment['Group']==i],longread_extract_output_per_flow_cell[longread_extract_output_per_flow_cell['Group']==i])
        else:
            (combined_summary_stats_df,longread_extract_minknow_version_dist,longread_extract_sample_rate_dist,longread_extract_flow_cells_per_experiment_dist)=longread_platform_qc_summary_statistics(longread_extract[longread_extract['Group'] == i],None,longread_extract_flow_cells_and_output_per_experiment[longread_extract_flow_cells_and_output_per_experiment['Group']==i],longread_extract_output_per_flow_cell[longread_extract_output_per_flow_cell['Group']==i])
        # write data frames with a row between each
        # write combined summary stats
        start_row = 0
        combined_summary_stats_df.to_excel(writer, startrow=start_row, index=False, sheet_name=i + ' statistics')
        # write flow cells per experiment distribution
        # add 1 to row number after combined_summary_stats_df exported
        start_row = start_row + len(combined_summary_stats_df) + 2
        longread_extract_flow_cells_per_experiment_dist.to_excel(writer, startrow=start_row, index=False, sheet_name=i + ' statistics')
        # write minknow version distribution
        start_row = start_row + len(longread_extract_flow_cells_per_experiment_dist) + 2
        longread_extract_minknow_version_dist.to_excel(writer, startrow=start_row, index=False, sheet_name=i + ' statistics')
        # write flow cells and output per flow cell on another worksheet
        longread_extract_output_per_flow_cell[longread_extract_output_per_flow_cell['Group']==i].to_excel(writer, index=False, sheet_name=i + ' output per FC')
        # write flow cells and output per unique experiment on another worksheet
        longread_extract_flow_cells_and_output_per_experiment[longread_extract_flow_cells_and_output_per_experiment['Group']==i].to_excel(writer, index=False, sheet_name=i + ' FC+output per expt')
        # eventually write joined platform QC/summary table to worksheet (to do)
    # close writer and save workbook
    writer.close()

# then add svg figures
# use openpyxl and pipe image data into new worksheets
# append new worksheets to existing workbook
workbook = openpyxl.load_workbook(results.output_file)

# probably functionalize all plots below and pass in group/group count variables
def make_report_plot_sequence(group_variable,legend_patches,user_palette,strip_plot_set):
    # use make_violinswarmplot_worksheet function to insert figures
    # plots that don't need cutoff included below to maintain order of spreadsheets (and order of figures)
    # if/else depending on whether -plot_cutoff set
    if results.plot_cutoff is True:
        # show topups in first four plots
        # test to check data type for first plotted variable: print(longread_extract['N50 (kb)'].dtype)
        make_violinswarmplot_worksheet(longread_extract,"N50 (kb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Read N50 plot',None,None,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Data output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Run data output plot',None,90,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Read Count (M)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Run read count plot',"Read Count (million reads)",None,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Starting Active Pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting active pores plot',"Starting active pores",6500,results.plot_title,True)
        # new fields to show
        # average active pores
        make_violinswarmplot_worksheet(longread_extract,"Average Active Pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Average active pores plot',"Average active pores",None,results.plot_title,True)
        # active pore AUC
        make_violinswarmplot_worksheet(longread_extract,"Active Pore AUC",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Active pore AUC plot',"Active pore AUC",None,results.plot_title,True)
        # starting pore occupancy
        make_violinswarmplot_worksheet(longread_extract,"Starting Pore Occupancy",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting pore occupancy plot',"Starting pore occupancy",None,results.plot_title,True)
        # average pore occupancy
        make_violinswarmplot_worksheet(longread_extract,"Average Pore Occupancy",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Average pore occupancy plot',"Average pore occupancy",None,results.plot_title,True)
        # hard coded Q score cutoffs of 8 based on fast basecalling
        # check if these values are even present using count (find how many non-NA values)
        if (longread_extract['Passed Modal Q Score'].count()>0):
            # passed modal Q score
            make_violinswarmplot_worksheet(longread_extract,"Passed Modal Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Passed modal Q score plot',"Passed modal Q score",8,results.plot_title,True)
        if (longread_extract['Failed Modal Q Score'].count()>0):
            # failed modal Q score
            make_violinswarmplot_worksheet(longread_extract,"Failed Modal Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Failed modal Q score plot',"Failed modal Q score",8,results.plot_title,True)
        if (longread_extract['Starting Median Translocation Speed'].count()>0):
            # starting translocation speed
            make_violinswarmplot_worksheet(longread_extract,"Starting Median Translocation Speed",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting transloc speed plot',"Starting translocation speed (bp/sec)",400,results.plot_title,True)
        if (longread_extract['Starting Median Q Score'].count()>0):
            # starting median Q score
            make_violinswarmplot_worksheet(longread_extract,"Starting Median Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting median Q score plot',"Starting median Q score",8,results.plot_title,True)
        if (longread_extract['Passed Bases (Gb)'].count()>0):
            # passed bases
            make_violinswarmplot_worksheet(longread_extract,"Passed Bases (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Passed bases plot',"Passed bases (Gb)",90,results.plot_title,True)
        if (longread_extract['Failed Bases (Gb)'].count()>0):
            # failed bases
            make_violinswarmplot_worksheet(longread_extract,"Failed Bases (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Failed bases plot',"Failed bases (Gb)",None,results.plot_title,True)
        if (longread_extract['Percentage Passed Bases'].count()>0):
            # percentage passed bases
            make_violinswarmplot_worksheet(longread_extract,"Percentage Passed Bases",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Percentage passed bases plot',"Percentage of bases passing filter",None,results.plot_title,True)
        # no topups in next three plots
        make_violinswarmplot_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Flow Cells",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Flow cells per experiment plot',"Flow cells",None,results.plot_title)
        make_violinswarmplot_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Total output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Output per experiment plot',None,90,results.plot_title)
        make_violinswarmplot_worksheet(longread_extract_output_per_flow_cell,"Flow cell output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Output per flow cell plot',None,90,results.plot_title)
        # platform qc violin/swarm plots
        if results.platform_qc is not None:
            # platform qc active pores 
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Platform QC active pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform QC active pores plot',None,6500,results.plot_title)
            # pore differences between sequencing and platform QC
            # show topups
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Pore Difference",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform-seq pore diff plot',"Platform to sequencing pore difference",None,results.plot_title,True)
            # time differences between sequencing and platform QC
            # show topups
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Time Difference",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform-seq time diff plot',"Platform to sequencing time difference",None,results.plot_title,True)
        # delivery date/batch/storage time swarmplots
        if results.delivery_date_batches is not None:
            # plot storage time in days
            make_violinswarmplot_worksheet(longread_extract,"Storage Time (Days)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Storage time plot',"Storage time (days)",None,results.plot_title,True)
                  
    else:
        # show topups in first four plots
        make_violinswarmplot_worksheet(longread_extract,"N50 (kb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Read N50 plot',None,None,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Data output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Run data output plot',None,None,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Read Count (M)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Run read count plot',"Read Count (million reads)",None,results.plot_title,True)
        make_violinswarmplot_worksheet(longread_extract,"Starting Active Pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting active pores plot',"Starting active pores",None,results.plot_title,True)
        # new fields to show
        # average active pores
        make_violinswarmplot_worksheet(longread_extract,"Average Active Pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Average active pores plot',"Average active pores",None,results.plot_title,True)
        # active pore AUC
        make_violinswarmplot_worksheet(longread_extract,"Active Pore AUC",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Active pore AUC plot',"Active pore AUC",None,results.plot_title,True)
        # starting pore occupancy
        make_violinswarmplot_worksheet(longread_extract,"Starting Pore Occupancy",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting pore occupancy plot',"Starting pore occupancy",None,results.plot_title,True)
        # average pore occupancy
        make_violinswarmplot_worksheet(longread_extract,"Average Pore Occupancy",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Average pore occupancy plot',"Average pore occupancy",None,results.plot_title,True)
        # hard coded Q score cutoffs of 8 based on fast basecalling
        # check if these values are even present using count (find how many non-NA values)
        if (longread_extract['Passed Modal Q Score'].count()>0):
            # passed modal Q score
            make_violinswarmplot_worksheet(longread_extract,"Passed Modal Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Passed modal Q score plot',"Passed modal Q score",None,results.plot_title,True)
        if (longread_extract['Failed Modal Q Score'].count()>0):
            # failed modal Q score
            make_violinswarmplot_worksheet(longread_extract,"Failed Modal Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Failed modal Q score plot',"Failed modal Q score",None,results.plot_title,True)
        if (longread_extract['Starting Median Translocation Speed'].count()>0):
            # starting translocation speed
            make_violinswarmplot_worksheet(longread_extract,"Starting Median Translocation Speed",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting transloc speed plot',"Starting translocation speed (bp/sec)",None,results.plot_title,True)
        if (longread_extract['Starting Median Q Score'].count()>0):
            # starting median Q score
            make_violinswarmplot_worksheet(longread_extract,"Starting Median Q Score",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Starting median Q score plot',"Starting median Q score",None,results.plot_title,True)
        if (longread_extract['Passed Bases (Gb)'].count()>0):
            # passed bases
            make_violinswarmplot_worksheet(longread_extract,"Passed Bases (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Passed bases plot',"Passed bases (Gb)",None,results.plot_title,True)
        if (longread_extract['Failed Bases (Gb)'].count()>0):
            # failed bases
            make_violinswarmplot_worksheet(longread_extract,"Failed Bases (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Failed bases plot',"Failed bases (Gb)",None,results.plot_title,True)
        if (longread_extract['Percentage Passed Bases'].count()>0):
            # percentage passed bases
            make_violinswarmplot_worksheet(longread_extract,"Percentage Passed Bases",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Percentage passed bases plot',"Percentage of bases passing filter",None,results.plot_title,True)
        # no topups in next three plots
        make_violinswarmplot_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Flow Cells",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Flow cells per experiment plot',"Flow cells",None,results.plot_title)
        make_violinswarmplot_worksheet(longread_extract_flow_cells_and_output_per_experiment,"Total output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Output per experiment plot',None,None,results.plot_title)
        make_violinswarmplot_worksheet(longread_extract_output_per_flow_cell,"Flow cell output (Gb)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Output per flow cell plot',None,None,results.plot_title)
        # platform qc violin/swarm plots
        if results.platform_qc is not None:
            # platform qc active pores 
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Platform QC active pores",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform QC active pores plot',None,None,results.plot_title)
            # pore differences between sequencing and platform QC
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Pore Difference",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform-seq pore diff plot',"Platform to sequencing pore difference",None,results.plot_title)
            # time differences between sequencing and platform QC
            make_violinswarmplot_worksheet(longread_extract_with_platform_qc_and_diff,"Time Difference",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Platform-seq time diff plot',"Platform to sequencing time difference",None,results.plot_title)
        # delivery date/batch/storage time swarmplots
        if results.delivery_date_batches is not None:
            # plot storage time in days
            make_violinswarmplot_worksheet(longread_extract,"Storage Time (Days)",group_variable,legend_patches,user_palette,strip_plot_set,workbook,'Storage time plot',"Storage time (days)",None,results.plot_title,True)
    # use make_active_pore_data_output_scatterplot function to add starting active pore vs. data output scatterplot
    # make_active_pore_data_output_scatterplot(longread_extract,workbook,'Active pores vs. data output',results.plot_title)
    # redo with make_scatterplot_worksheet
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Active pores vs. data output",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
    # do also with read count
    # cutoffs not known yet for read count
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Active pores vs. read count",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='Read Count (M)',prop_point_size=False,size_column=None)            
    # make_active_pore_flow_cell_output_scatterplot(longread_extract_output_per_flow_cell,workbook,'Active pores vs. flow cell output',results.plot_title)
    # make_active_pore_read_n50_scatterplot(longread_extract,workbook,'Active pores vs. read N50',results.plot_title)
    # redo with make_scatterplot_worksheet
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Active pores vs. read N50",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='N50 (kb)',prop_point_size=False,size_column=None)            
    # make_read_n50_data_output_scatterplot(longread_extract,workbook,'Read N50 vs. data output',results.plot_title)
    # redo with make_scatterplot_worksheet
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Read N50 vs. data output",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=True,x_variable='N50 (kb)',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
    # add read N50 vs. data output without regression line
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Read N50 vs. data no line",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='N50 (kb)',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)     
    # do also with read count
    # redo with make_scatterplot_worksheet
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Read N50 vs. read count",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=True,x_variable='N50 (kb)',y_variable='Read Count (M)',prop_point_size=False,size_column=None)            
    # add read N50 vs. data output without regression line
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Read N50 vs. count no line",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='N50 (kb)',y_variable='Read Count (M)',prop_point_size=False,size_column=None)     
    # add starting active pores vs. average active pores
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Start vs. avg active pores",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='Average Active Pores',prop_point_size=False,size_column=None)            
    # add starting active pores vs. active pore AUC
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Start active pores vs. AUC",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='Active Pore AUC',prop_point_size=False,size_column=None)            
    # add average active pores vs. data output
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Avg active pores vs. data",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Average Active Pores',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
    # add active pore AUC vs. data output
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Active pore AUC vs. data",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Active Pore AUC',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
    # add active pore AUC vs. read count
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore AUC vs. read count",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Active Pore AUC',y_variable='Read Count (M)',prop_point_size=False,size_column=None)            
    
    # add average pore occupancy vs. data output
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Avg pore occup. vs. data",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Average Pore Occupancy',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
    # add average median translocation speed vs. average median Q score if both columns have data present
    if ((longread_extract['Average Median Translocation Speed Over Time'].count()>0) and (longread_extract['Average Median Q Score Over Time'].count()>0)):
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Transloc. speed vs. Q score",title=results.plot_title,x_cutoffs=[400],x_cutoff_colors=['blue'],y_cutoffs=[8],y_cutoff_colors=['green'],show_run_colors=True,show_reg_line=False,x_variable='Average Median Translocation Speed Over Time',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
    # add data output vs. average median Q score if Q score present
    if (longread_extract['Average Median Q Score Over Time'].count()>0):
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Data output vs. Q score",title=results.plot_title,x_cutoffs=[90],x_cutoff_colors=['gray'],y_cutoffs=[8],y_cutoff_colors=['green'],show_run_colors=True,show_reg_line=False,x_variable='Data output (Gb)',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
    # add starting active pores vs. average median Q score if Q score present
    if (longread_extract['Average Median Q Score Over Time'].count()>0):
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Start. pores vs. Q score",title=results.plot_title,x_cutoffs=[5000,6500],x_cutoff_colors=['red','green'],y_cutoffs=[8],y_cutoff_colors=['blue'],show_run_colors=True,show_reg_line=False,x_variable='Starting Active Pores',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
    # add active pore AUC vs. average median Q score if Q score present
    if (longread_extract['Average Median Q Score Over Time'].count()>0):
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Active pore AUC vs. Q score",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[8],y_cutoff_colors=['green'],show_run_colors=True,show_reg_line=False,x_variable='Active Pore AUC',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
    # add passed bases vs. passed modal q score if both columns have data present
    if ((longread_extract['Passed Bases (Gb)'].count()>0) and (longread_extract['Passed Modal Q Score'].count()>0)):
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Passed bases vs. Q score",title=results.plot_title,x_cutoffs=[90],x_cutoff_colors=['gray'],y_cutoffs=[8],y_cutoff_colors=['green'],show_run_colors=True,show_reg_line=False,x_variable='Passed Bases (Gb)',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
    # log10 transform platform qc active pores, starting active pores, and pore differences
    if results.platform_qc is not None:
        longread_extract_with_platform_qc_and_diff['log_platform_qc_active_pores']=np.log10(pd.to_numeric(longread_extract_with_platform_qc_and_diff['Platform QC active pores']))
        longread_extract_with_platform_qc_and_diff['log_starting_active_pores']=np.log10(pd.to_numeric(longread_extract_with_platform_qc_and_diff['Starting Active Pores']))
        longread_extract_with_platform_qc_and_diff['log_pore_difference']=np.log10(pd.to_numeric(longread_extract_with_platform_qc_and_diff['Pore Difference']))
        # plots to include if platform qc provided
        # pore difference violin/swarmplot
        # pore measurement time difference violin/swarmplot
        # all done above
        # platform qc active pores vs. starting active pores
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pqc vs. starting active pores",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Platform QC active pores',y_variable='Starting Active Pores',prop_point_size=False,size_column=None)        
        # platform qc active pores vs. pore difference
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pqc vs. pore difference",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Platform QC active pores',y_variable='Pore Difference',prop_point_size=False,size_column=None)    
        # pore difference vs. time difference scatterplot with run colors set
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore vs. time difference",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Time Difference',y_variable='Pore Difference',prop_point_size=False,size_column=None)    
        # pore difference vs. run output scatterplot
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore difference vs. run output",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Pore Difference',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)        
        # add pore difference vs. active pore auc scatterplot
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore diff. vs. active pore AUC",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Pore Difference',y_variable='Active Pore AUC',prop_point_size=False,size_column=None)        
        # add pore difference vs. average median Q score if Q score present
        if (longread_extract['Average Median Q Score Over Time'].count()>0):
            make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore diff. vs. Q score",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[8],y_cutoff_colors=['blue'],show_run_colors=True,show_reg_line=False,x_variable='Pore Difference',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)        
        # add pore difference vs. pore occupancy scatterplot
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore occup. vs. pore diff.",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Starting Pore Occupancy',y_variable='Pore Difference',prop_point_size=False,size_column=None)        
        # read N50 vs. run output scatterplot with coloring by run type and size based on pore difference
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"N50 output pore diff",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='N50 (kb)',y_variable='Data output (Gb)',prop_point_size=True,size_column='Pore Difference')    
        # read N50 vs. run output scatterplot with coloring by run type and size based on starting active pores
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"N50 output starting pores",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='N50 (kb)',y_variable='Data output (Gb)',prop_point_size=True,size_column='Starting Active Pores')
        # run matched platform qc active pores over time
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Platform QC pores over time",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[6500],y_cutoff_colors=['green'],show_run_colors=True,show_reg_line=False,x_variable='Platform QC date',y_variable='Platform QC active pores',prop_point_size=False,size_column=None,has_date_time=True)
        # pore difference over time
        make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore difference over time",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Run date',y_variable='Pore Difference',prop_point_size=False,size_column=None,has_date_time=True)    
        # delivery date/batch/storage time swarmplots
        if results.delivery_date_batches is not None:
            # storage time vs. pore difference
            make_scatterplot_worksheet(longread_extract_with_platform_qc_and_diff,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Pore diff. vs storage",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Storage Time (Days)',y_variable='Pore Difference',prop_point_size=False,size_column=None,has_date_time=True)    
            
    # run output over time
    make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Run output over time",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Run date',y_variable='Data output (Gb)',prop_point_size=False,size_column=None,has_date_time=True)    
    # delivery date/batch/storage time swarmplots
    if results.delivery_date_batches is not None:
        # storage time vs. data output
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Storage vs. data output",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[90],y_cutoff_colors=['gray'],show_run_colors=True,show_reg_line=False,x_variable='Storage Time (Days)',y_variable='Data output (Gb)',prop_point_size=False,size_column=None)            
        # storage time vs. starting active pores
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Storage vs. starting pores",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[5000,6500],y_cutoff_colors=['red','green'],show_run_colors=True,show_reg_line=False,x_variable='Storage Time (Days)',y_variable='Starting Active Pores',prop_point_size=False,size_column=None)            
        # storage time vs. active pore AUC
        make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Storage vs. pore AUC",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=None,y_cutoff_colors=None,show_run_colors=True,show_reg_line=False,x_variable='Storage Time (Days)',y_variable='Active Pore AUC',prop_point_size=False,size_column=None)            
        # storage time vs. average median Q score over run
        if (longread_extract['Average Median Q Score Over Time'].count()>0):
            make_scatterplot_worksheet(longread_extract,group_variable,legend_patches,user_palette,strip_plot_set,workbook,"Storage vs. Q score",title=results.plot_title,x_cutoffs=None,x_cutoff_colors=None,y_cutoffs=[8],y_cutoff_colors=['blue'],show_run_colors=True,show_reg_line=False,x_variable='Storage Time (Days)',y_variable='Average Median Q Score Over Time',prop_point_size=False,size_column=None)            
            
    
# run through plots depending on whether group variable and group count variables set
if grouped is False:
    make_report_plot_sequence(None,legend_patches,results.colors,results.strip_plot)
# if group variable set
elif grouped is True:
    # show group count if variable set
    if results.show_group_count is True:
        make_report_plot_sequence('Group and count',legend_patches,results.colors,results.strip_plot)
    else:
        make_report_plot_sequence('Group',legend_patches,results.colors,results.strip_plot)

# save workbook when done
workbook.save(results.output_file)

# script complete
quit()